from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill

from datetime import datetime

redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
blackFill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')

WAS_STATUS_MESSAGES={
    'SHUTDOWN':{'color':redFill,'message':'서비스 꺼짐'}
    ,'STANDBY':{'color':redFill,'message':'Application deploy 실패'}
    ,'FAILED':{'color':redFill,'message':'서비스 불가'}
    ,'NOTCHECKED':{'color':redFill,'message':'모니터링 불가'}
    ,'UNCHECKED':{'color':blackFill,'message':'Agent 이상'}
    }

def _getCellByInstance(ws, start_row, end_row, start_col, end_col, group_column, group_id, value):

    for row in range(start_row, end_row + 1):
        for cell in range(start_col, end_col + 1):
            if ws[row][cell].value == value and ws[row][group_column].value == group_id:
                return ws[row][cell]

    return None

def createDailyCheckXlsx(master_file_name, target_file_name, result, now):

    ymd = now.strftime("%Y.%m.%d")

    wb = load_workbook(master_file_name)
    ws = wb.worksheets[0]
    ws['F2'] = ymd
    t_message_l = []
    rtn = 1

    for r in result:
        for ri in r['instances']:
            cell = _getCellByInstance(ws, 11, 94, 1, 13, 16, r['was_id'], ri['was_instance_id'])
            if cell:
                colFill = WAS_STATUS_MESSAGES[ri['status']]['color']
                message = ri['was_instance_id'] + ':' + WAS_STATUS_MESSAGES[ri['status']]['message']
                t_message_l.append(message)
                cell.fill = colFill
                if ws[cell.row][14].value:
                    ws[cell.row][14].value = ws[cell.row][14].value + ', ' + message
                else:
                    ws[cell.row][14].value = message

                #print(cell.row, message)

    try:
        wb.save(filename=target_file_name)
    except PermissionError as e:
        rtn = -1

    t_message = ', '.join(t_message_l)

    return rtn, t_message

if __name__ == "__main__":
    
    ymd = datetime.now().strftime("%Y.%m.%d")
    master_file_name = 'C:\\Users\\KDB\\Downloads\\미들웨어일일점검_template.xlsx'
    target_file_name = 'C:\\Users\\KDB\\Downloads\\미들웨어일일점검_'+ymd+'.xlsx'
    #result = [{'was_id': 'PAPM_Domain', 'instances': [{'was_instance_id': 'GUI_MS21', 'status': 'SHUTDOWN'}]}, {'was_id': 'PICI_Domain', 'instances': [{'was_instance_id': 'IMG_MS21', 'status': 'SHUTDOWN'}]}, {'was_id': 'PADW_Domain', 'instances': [{'was_instance_id': 'RAY_MS21', 'status': 'SHUTDOWN'}]}, {'was_id': 'PHMP_Domain', 'instances': [{'was_instance_id': 'WLA_MS31', 'status': 'SHUTDOWN'}, {'was_instance_id': 'SOA_MS21', 'status': 'SHUTDOWN'}, {'was_instance_id': 'SOA_MS31', 'status': 'SHUTDOWN'}, {'was_instance_id': 'WLA_MS21', 'status': 'SHUTDOWN'}]}, {'was_id': 'PWMM_Domain', 'instances': [{'was_instance_id': 'WMM_MS23', 'status': 'SHUTDOWN'}]}]
    result = [{'was_id': 'PCBK_Domain', 'instances': [{'was_instance_id': 'ONL_MS28', 'status': 'SHUTDOWN'}]}, {'was_id': 'PICI_Domain', 'instances': [{'was_instance_id': 'ICI_MS14', 'status': 'SHUTDOWN'}]}, {'was_id': 'PADW_Domain', 'instances': [{'was_instance_id': 'MKR_MST1', 'status': 'STANDBY'}]}, {'was_id': 'PHMP_Domain', 'instances': [{'was_instance_id': 'APR_MS21', 'status': 'NOTCHECKED'}, {'was_instance_id': 'CHP_MS23', 'status': 'SHUTDOWN'}, {'was_instance_id': 'CBP_MS32', 'status': 'SHUTDOWN'}, {'was_instance_id': 'WLA_MS21', 'status': 'SHUTDOWN'}]}, {'was_id': 'PWMM_Domain', 'instances': [{'was_instance_id': 'WMM_MS23', 'status': 'SHUTDOWN'}]}]

    rtn, t_message = createDailyCheckXlsx(master_file_name, target_file_name, result, datetime.now())
    print(rtn, t_message)
